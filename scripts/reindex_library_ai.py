from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

from dotenv import load_dotenv
import httpx
from openpyxl import load_workbook
from pinecone import Pinecone
from pypdf import PdfReader


DEFAULT_EXCEL_PATH = Path("data/list coventry.xlsx")
DEFAULT_CLEAN_CSV_PATH = Path("data/library_catalog_clean.csv")
DEFAULT_PDF_DIR = Path("pdf")
DEFAULT_TEMPLATE_DIR = Path("app/templates")
EMBED_TEXT_LIMIT = 250


@dataclass(frozen=True)
class BookRecord:
    author: str
    title: str
    classification_number: str
    source_row: int


@dataclass(frozen=True)
class VectorRecord:
    id: str
    text: str
    metadata: dict[str, Any]


class TemplateTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._parts: list[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag in {"script", "style"}:
            self._skip_depth += 1
            return

        attr_map = {name.lower(): value for name, value in attrs if value}
        if tag == "a" and attr_map.get("href"):
            self._parts.append(f"Link: {attr_map['href']}")
        if tag == "img" and attr_map.get("alt"):
            self._parts.append(attr_map["alt"])
        if tag in {"p", "div", "section", "article", "li", "tr", "br", "h1", "h2", "h3", "h4", "h5"}:
            self._parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"script", "style"} and self._skip_depth:
            self._skip_depth -= 1
            return
        if tag in {"p", "div", "section", "article", "li", "tr", "h1", "h2", "h3", "h4", "h5"}:
            self._parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self._skip_depth:
            self._parts.append(data)

    @property
    def text(self) -> str:
        return normalize_text(" ".join(self._parts))


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = html.unescape(text)
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text, flags=re.UNICODE)
    return text.strip()


def clean_title(value: Any) -> str:
    text = normalize_text(value)
    return text.rstrip(" /:;")


def clean_classification(value: Any) -> str:
    text = normalize_text(value)
    return re.sub(r"\s+", " ", text).strip()


def stable_id(prefix: str, *parts: str) -> str:
    raw = "|".join(parts)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:20]
    return f"{prefix}:{digest}"


def load_books(
    excel_path: Path,
    *,
    keep_missing_classification: bool = False,
    keep_title_author_equals: bool = False,
) -> tuple[list[BookRecord], dict[str, int]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    headers = [normalize_text(cell) for cell in next(rows)]
    header_index = {header: idx for idx, header in enumerate(headers) if header}

    required = {"Personal_name", "Title"}
    missing = required - set(header_index)
    if missing:
        raise ValueError(f"Missing required Excel columns: {', '.join(sorted(missing))}")

    stats = {
        "source_rows": 0,
        "missing_classification_skipped": 0,
        "title_author_equal_skipped": 0,
        "empty_title_author_skipped": 0,
        "duplicates_skipped": 0,
    }
    books: list[BookRecord] = []
    seen: set[str] = set()

    for source_row, row in enumerate(rows, start=2):
        stats["source_rows"] += 1

        def field(name: str) -> Any:
            idx = header_index.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        author = normalize_text(field("Personal_name"))
        title = clean_title(field("Title"))
        classification = clean_classification(field("Classification_number"))
        if not classification:
            classification = clean_classification(field("Item_number"))

        if not author and not title:
            stats["empty_title_author_skipped"] += 1
            continue
        if not classification and not keep_missing_classification:
            stats["missing_classification_skipped"] += 1
            continue
        if (
            not keep_title_author_equals
            and author
            and title
            and author.casefold().strip(". ") == title.casefold().strip(". ")
        ):
            stats["title_author_equal_skipped"] += 1
            continue

        dedupe_key = "\u241f".join(
            [author.casefold(), title.casefold(), classification.casefold()]
        )
        if dedupe_key in seen:
            stats["duplicates_skipped"] += 1
            continue
        seen.add(dedupe_key)
        books.append(
            BookRecord(
                author=author,
                title=title,
                classification_number=classification,
                source_row=source_row,
            )
        )

    return books, stats


def write_clean_catalog_csv(books: list[BookRecord], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["author", "title", "classification_number", "source_row"],
        )
        writer.writeheader()
        for book in books:
            writer.writerow(
                {
                    "author": book.author,
                    "title": book.title,
                    "classification_number": book.classification_number,
                    "source_row": book.source_row,
                }
            )


def strip_jinja(text: str) -> str:
    text = re.sub(r"\{#.*?#\}", " ", text, flags=re.DOTALL)
    text = re.sub(r"\{%.*?%\}", " ", text, flags=re.DOTALL)
    text = re.sub(r"\{\{.*?\}\}", " ", text, flags=re.DOTALL)
    return text


def extract_template_text(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    parser = TemplateTextExtractor()
    parser.feed(strip_jinja(raw))
    return parser.text


def extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    pages: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        text = normalize_text(text)
        if text:
            pages.append(f"Page {index}: {text}")
    return "\n\n".join(pages)


def chunk_text(text: str, chunk_size: int, overlap: int) -> list[str]:
    text = normalize_text(text)
    if not text:
        return []
    chunks: list[str] = []
    start = 0
    text_length = len(text)
    while start < text_length:
        end = min(start + chunk_size, text_length)
        if end < text_length:
            split_at = text.rfind(" ", start, end)
            if split_at > start + int(chunk_size * 0.6):
                end = split_at
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= text_length:
            break
        next_start = max(0, end - overlap)
        if next_start <= start:
            next_start = end
        start = next_start
    return chunks


def build_book_vector_records(books: list[BookRecord]) -> list[VectorRecord]:
    records: list[VectorRecord] = []
    for book in books:
        title = book.title or "Untitled"
        author = book.author or "Unknown"
        text = (
            "Book catalog record. "
            f"Title / Название / Атауы: {title}. "
            f"Author / Автор: {author}. "
            "Classification number / Shelf number / Номер классификации / Полочный номер: "
            f"{book.classification_number}."
        )
        record_id = stable_id(
            "book",
            book.author,
            book.title,
            book.classification_number,
        )
        records.append(
            VectorRecord(
                id=record_id,
                text=text,
                metadata={
                    "source": "book_catalog",
                    "source_type": "book",
                    "title": title,
                    "author": author,
                    "classification_number": book.classification_number,
                    "source_row": book.source_row,
                    "text_preview": text,
                },
            )
        )
    return records


def build_pdf_vector_records(pdf_dir: Path, chunk_size: int, overlap: int) -> list[VectorRecord]:
    records: list[VectorRecord] = []
    for path in sorted(pdf_dir.glob("*.pdf")):
        text = extract_pdf_text(path)
        for chunk_index, chunk in enumerate(chunk_text(text, chunk_size, overlap), start=1):
            preview = f"Document: {path.name}. {chunk}"
            records.append(
                VectorRecord(
                    id=stable_id("pdf", path.as_posix(), str(chunk_index), chunk[:80]),
                    text=preview,
                    metadata={
                        "source": path.as_posix(),
                        "source_type": "pdf",
                        "filename": path.name,
                        "title": path.stem.replace("_", " "),
                        "chunk": chunk_index,
                        "text_preview": preview[:35000],
                    },
                )
            )
    return records


def build_template_vector_records(template_dir: Path, chunk_size: int, overlap: int) -> list[VectorRecord]:
    records: list[VectorRecord] = []
    for path in sorted(template_dir.rglob("*.html")):
        text = extract_template_text(path)
        for chunk_index, chunk in enumerate(chunk_text(text, chunk_size, overlap), start=1):
            preview = f"Website information from {path.as_posix()}. {chunk}"
            records.append(
                VectorRecord(
                    id=stable_id("template", path.as_posix(), str(chunk_index), chunk[:80]),
                    text=preview,
                    metadata={
                        "source": path.as_posix(),
                        "source_type": "website",
                        "filename": path.name,
                        "title": path.stem,
                        "chunk": chunk_index,
                        "text_preview": preview[:35000],
                    },
                )
            )
    return records


def batches(items: list[VectorRecord], batch_size: int) -> Iterable[list[VectorRecord]]:
    for start in range(0, len(items), batch_size):
        yield items[start : start + batch_size]


def embed_text(client: httpx.Client, ollama_url: str, model: str, text: str) -> list[float]:
    response = client.post(
        f"{ollama_url.rstrip('/')}/api/embeddings",
        json={"model": model, "prompt": text[:EMBED_TEXT_LIMIT]},
    )
    if response.status_code >= 400:
        raise RuntimeError(
            f"Ollama embeddings failed with HTTP {response.status_code}: "
            f"{response.text[:500]}"
        )
    data = response.json()
    embedding = data.get("embedding")
    if not isinstance(embedding, list) or not embedding:
        raise RuntimeError(f"Ollama returned no embedding for model {model!r}")
    return embedding


def pinecone_namespace_kwargs(namespace: str) -> dict[str, str]:
    namespace = namespace.strip()
    return {"namespace": namespace} if namespace else {}


def to_plain(value: Any) -> Any:
    if hasattr(value, "to_dict"):
        return value.to_dict()
    if isinstance(value, dict):
        return {key: to_plain(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [to_plain(item) for item in value]
    return value


def upsert_records(
    records: list[VectorRecord],
    *,
    pinecone_api_key: str,
    pinecone_index: str,
    namespace: str,
    ollama_url: str,
    embed_model: str,
    batch_size: int,
    clear: bool,
) -> None:
    pc = Pinecone(api_key=pinecone_api_key)
    index = pc.Index(pinecone_index)
    namespace_kwargs = pinecone_namespace_kwargs(namespace)
    prepared_vectors = []

    print(f"Preparing embeddings for {len(records)} records before touching Pinecone...")
    with httpx.Client(timeout=60.0) as client:
        for record_number, record in enumerate(records, start=1):
            vector = embed_text(client, ollama_url, embed_model, record.text)
            prepared_vectors.append(
                {
                    "id": record.id,
                    "values": vector,
                    "metadata": {
                        key: value
                        for key, value in record.metadata.items()
                        if value is not None and value != ""
                    },
                }
            )
            if record_number == 1 or record_number % 100 == 0 or record_number == len(records):
                print(f"Prepared embeddings: {record_number}/{len(records)}")

    try:
        before_stats = to_plain(index.describe_index_stats())
        print("Pinecone stats before:")
        print(json.dumps(before_stats, ensure_ascii=False, indent=2, default=str))
    except Exception as exc:
        print(f"Could not read Pinecone stats before upsert: {exc}")

    if clear:
        print(f"Deleting all records in namespace {namespace or '<default>'}...")
        index.delete(delete_all=True, **namespace_kwargs)
        time.sleep(5)

    total = 0
    for start in range(0, len(prepared_vectors), batch_size):
        batch_number = start // batch_size + 1
        vectors = prepared_vectors[start : start + batch_size]
        index.upsert(vectors=vectors, **namespace_kwargs)
        total += len(vectors)
        print(f"Upserted batch {batch_number}: {len(vectors)} records, total {total}")

    try:
        after_stats = to_plain(index.describe_index_stats())
        print("Pinecone stats after:")
        print(json.dumps(after_stats, ensure_ascii=False, indent=2, default=str))
    except Exception as exc:
        print(f"Could not read Pinecone stats after upsert: {exc}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Clean the Coventry library catalog and reindex AI data into Pinecone."
    )
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--clean-csv", type=Path, default=DEFAULT_CLEAN_CSV_PATH)
    parser.add_argument("--pdf-dir", type=Path, default=DEFAULT_PDF_DIR)
    parser.add_argument("--template-dir", type=Path, default=DEFAULT_TEMPLATE_DIR)
    parser.add_argument("--chunk-size", type=int, default=1200)
    parser.add_argument("--chunk-overlap", type=int, default=150)
    parser.add_argument("--batch-size", type=int, default=50)
    parser.add_argument("--limit", type=int, default=0, help="Limit vector records for testing.")
    parser.add_argument("--apply", action="store_true", help="Write vectors to Pinecone.")
    parser.add_argument("--clear", action="store_true", help="Delete all records in the target namespace first.")
    parser.add_argument("--keep-missing-classification", action="store_true")
    parser.add_argument("--keep-title-author-equals", action="store_true")
    parser.add_argument("--namespace", default=os.getenv("PINECONE_NAMESPACE", "__default__"))
    return parser.parse_args()


def main() -> int:
    load_dotenv(dotenv_path=Path(".env"))
    args = parse_args()

    books, book_stats = load_books(
        args.excel,
        keep_missing_classification=args.keep_missing_classification,
        keep_title_author_equals=args.keep_title_author_equals,
    )
    write_clean_catalog_csv(books, args.clean_csv)

    vector_records = []
    vector_records.extend(build_book_vector_records(books))
    vector_records.extend(build_pdf_vector_records(args.pdf_dir, args.chunk_size, args.chunk_overlap))
    vector_records.extend(
        build_template_vector_records(args.template_dir, args.chunk_size, args.chunk_overlap)
    )
    if args.limit:
        vector_records = vector_records[: args.limit]

    source_counts: dict[str, int] = {}
    for record in vector_records:
        source_type = str(record.metadata.get("source_type", "unknown"))
        source_counts[source_type] = source_counts.get(source_type, 0) + 1

    print("Catalog cleanup:")
    print(json.dumps(book_stats, ensure_ascii=False, indent=2))
    print(f"Clean catalog rows: {len(books)}")
    print(f"Clean CSV: {args.clean_csv}")
    print("Vector record counts:")
    print(json.dumps(source_counts, ensure_ascii=False, indent=2))
    print(f"Total vector records: {len(vector_records)}")

    if not args.apply:
        print("Dry run only. Re-run with --apply to upsert to Pinecone.")
        return 0

    pinecone_api_key = os.getenv("PINECONE_API_KEY", "")
    pinecone_index = os.getenv("PINECONE_INDEX", "library-assistant")
    ollama_url = os.getenv("OLLAMA_URL", "http://localhost:11434")
    embed_model = os.getenv("EMBED_MODEL", "all-minilm")

    if not pinecone_api_key:
        print("PINECONE_API_KEY is missing in .env", file=sys.stderr)
        return 2
    if not pinecone_index:
        print("PINECONE_INDEX is missing in .env", file=sys.stderr)
        return 2

    upsert_records(
        vector_records,
        pinecone_api_key=pinecone_api_key,
        pinecone_index=pinecone_index,
        namespace=args.namespace,
        ollama_url=ollama_url,
        embed_model=embed_model,
        batch_size=args.batch_size,
        clear=args.clear,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
